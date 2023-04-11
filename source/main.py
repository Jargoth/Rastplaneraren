from tkinter import *
from tkinter import ttk, colorchooser, messagebox, filedialog
import xml.dom.minidom
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import codecs
import datetime


# Project modules
from settings import getsettings, xml_new_task, xml_add_person, xml_delete_task, xml_save_excel_template, xml_save_excel
from settings import delete_announcement
from plan_breaks import plan_breaks
import log_system

log, logfile = logging.start()

# logging
if log['start_stop']:
    time = datetime.datetime.now()
    with open(logfile, 'a') as f:
        f.write(f'{time.hour}:{time.minute}:{time.second} start_stop: program started\n')

version = '0.1.4'


def button_color(row, col):
    # This function is called when clicking on section of the schedule.
    # It changes that section to the color representing the task in activetask
    # if the section already is the selected colour the closest is chosen instead

    # logging
    if log['change_task']:
        time = datetime.datetime.now()
        task_name = tasksvariable[activetask.get()][1]
        task_name_from = tasksvariable[person[row][4][col][1]][1]
        temp = f'{str((int(col/4))+8)}:{str((col%4)*15)}'
        with open(logfile, 'a') as f:
            f.write(
                f'{time.hour}:{time.minute}:{time.second} change_task: row: {row} time: {temp} to_task: {task_name} from_task: {task_name_from}\n')

    numtries = 1  # number of tries to select the closest colour thats not activetask if the selected is the same
    forward = False

    # if selected section is the same as selected task
    if person[row][4][col][1] == activetask.get():
        completed = False
        while not completed:
            if forward:  # check forward in the schedule

                # still the same so turn the other way and ad 1 step
                if person[row][4][col + numtries][1] == activetask.get():
                    # still the same so turn the other way and ad 1 step
                    forward = False
                    numtries = numtries + 1

                else:  # Found the color to change to
                    oldtask = person[row][4][col + numtries][1]
                    person[row][4][col][0]['bg'] = f'#{tasksvariable[oldtask][2]}'
                    person[row][4][col][1] = oldtask
                    completed = True

                    #logging
                    if log['change_task']:
                        time = datetime.datetime.now()
                        task_name = tasksvariable[oldtask][1]
                        with open(logfile, 'a') as f:
                            f.write(
                                f'{time.hour}:{time.minute}:{time.second} change_task: row: {row} time: {temp} change_to: {task_name}\n')

            else: #check backwards in the schedule
                if person[row][4][col - numtries][1] == activetask.get():  # still the same so turn the other way
                    forward = True
                else:  # Found the color to change to
                    oldtask = person[row][4][col - numtries][1]
                    person[row][4][col][0]['bg'] = f'#{tasksvariable[oldtask][2]}'
                    person[row][4][col][1] = oldtask
                    completed = True

    # if the selected section differs from active task. Change colour
    else:
        person[row][4][col][0]['bg'] = f'#{tasksvariable[activetask.get()][2]}'
        person[row][4][col][1] = activetask.get()


def add_time(row, workinghours, type):
    # This function is run when entering starting and stoping working hours on todays schedule
    # It shows buttons that corresponds to the correct quarter of an hour.
    # The buttons will be set to a color that curresponds to the default color of that employee.

    # logging
    if log['add_time']:
        time = datetime.datetime.now()
        with open(logfile, 'a') as f:
            f.write(
            f'{time.hour}:{time.minute}:{time.second} add_time: row: {row} workinghours: {workinghours} type: {type}\n')

    #check if the format is correct
    if type == 'key':

        # Checks if value is a number or - or :
        valid = '0123456789:-'
        for char in workinghours:
            if char not in valid:
                return False

        # splits on - if - is present
        if '-' in workinghours:
            start_end = workinghours.split('-')
        else:
            start_end = [workinghours]

        # splits on : if : is presnt
        for se in start_end:
            if ':' in se:
                start = se.split(':')
            else:
                start = [se]
            # returns False if more than 2 numbers is entered
            for s in start:
                if len(s) >= 3:
                    return False

    # prints out when leaving
    elif type == 'focusout':
        try:

            # start by clearing all the buttons
            for i in range(52):
                person[int(row)][4][i][1] = -1  # the index number of the currents task is set to -1 (= no task)
                person[int(row)][4][i][0].grid_remove()  # don't show the button

            # extracting start and stop time
            startend = workinghours.split('-')
            start = startend[0].split(':')
            end = startend[1].split(':')
            starthour = int((int(start[0]) - 8) * 4)
            startmin = int(int(start[1]) / 15)
            if starthour < 0:
                starthour = 0
                startmin = 0
            endhour = int((int(end[0]) - 8) * 4)
            endmin = int(int(end[1]) / 15)
            curr = starthour + startmin
            if curr < 0:
                curr = 0

            # set the task index, set correct color, show the buttons
            for i in range(endhour + endmin - starthour - startmin):
                if i + curr < 52:
                    person[int(row)][4][i + curr][1] = person[int(row)][5][1]
                    person[int(row)][4][i + curr][0]['bg'] = f"#{tasksvariable[person[int(row)][5][1]][2]}"
                    person[int(row)][4][i + curr][0].grid()

        except:
            return False
    return True


def add_person(row, name, type):
    # This function is runned when you enter an employee name on todays schedule
    # It connects that name to correct employee to get the special settings for him/her
    # If a matching employee isn't found it adds a new with default settings

    # check if the format is correct
    if type == 'key':

        # Checks if value is a number or - or :
        valid = 'abcdefghijklmnopqrstuvwxyzåäöABCDEFGHIJKLMNOPQRSTUVWXYZÅÄÖ üÜ'
        for char in name:
            if char not in valid:
                return False

    # check if there's a matching employee
    elif type == 'focusout':
        if name:
            # logging
            if log['add_name']:
                time = datetime.datetime.now()
                with open(logfile, 'a') as f:
                    f.write(f'{time.hour}:{time.minute}:{time.second} add_name: row: {row} name: {name}\n')

            name = name.lower().capitalize()
            person_id = -1  # -1 means there's no matching
            i = 0
            for (i, employee) in enumerate(employees):
                if name == employee[0]:
                    person_id = i

            # If there's a new employee
            if person_id == -1:
                person_id = xml_add_person(name, tasksvariable, employees, i)

            # Set all variables to match the employee
            default_task = employees[person_id][2]
            task_color = tasksvariable[0][2]
            default_task_number = 0
            for tn, t in enumerate(tasksvariable):
                if t[0] == default_task:
                    task_color = t[2]
                    default_task_number = tn
            print(row)
            person[int(row)][5][0]['bg'] = f'#{task_color}'
            person[int(row)][5][1] = default_task_number

            add_row()

    return True


def move_settings_window(e):
    # This function is called when moving the settings window
    # It moves the window to the possition of the mouse

    settingsWindow.geometry(f'+{e.x_root}+{e.y_root}')


def show_new_task():
    # Show the widgets needed to add a new task

    # set default values
    newtaskwidgets[0]['text'] = 'Lägg till en ny uppgift'
    newtaskname.set('Ny uppgift')
    newtaskcolor.set('555555')
    newtaskwidgets[4]['bg'] = '#555555'
    newtaskauto_generate.set(False)
    newtaskdefault_certified.set(False)
    newtaskschedule_length.set('60')
    newtaskschedule_max_times.set('3')

    # show everything except edit task save
    for newtaskwidgetn, newtaskwidget in enumerate(newtaskwidgets):
        if newtaskwidgetn != 15 and newtaskwidgetn != 16:
            newtaskwidget.grid()


def hide_new_task():
    # Hide the widgets to add a new task

    for newtaskwidgetn, newtaskwidget in enumerate(newtaskwidgets):
        if newtaskwidgetn == 15 or newtaskwidgetn == 16:  # edit task save and delete
            for newtaskwidge in newtaskwidget:
                newtaskwidge.grid_remove()
        else:  # everything else
            newtaskwidget.grid_remove()


def select_color_new_task():
    # Select color for a new task. It opens a colorchooser window and puts the return value into newtaskwidgets[4]['bg']

    colorwindow = colorchooser.askcolor(initialcolor=f'#{newtaskcolor.get()}', parent=root)
    newtaskcolor.set(colorwindow[1][1:])
    newtaskwidgets[4]['bg'] = f'#{colorwindow[1][1:]}'


def new_task_save():
    # Saves the new task. It updates the XML and all associated variables

    # Update the XML
    xml_new_task(tasksvariable, newtaskname, newtaskcolor, newtaskauto_generate, newtaskdefault_certified, newtaskschedule_length, newtaskschedule_max_times)

    # Update activetasks radiobuttons on todays schedulue
    i = len(tasksvariable) - 1
    ttk.Radiobutton(topframe, text=str(newtaskname.get()), variable=activetask, value=i).grid(column=i, row=0)
    Button(tasksframe,
           text=str(newtaskcolor.get()),
           bg=f'#{newtaskcolor.get()}').grid(row=i, column=1, padx=2, pady=2)
    label = Label(tasksframe, text=str(newtaskname.get()))
    label.grid(row=i, column=0, padx=2, pady=2)
    label.bind('<ButtonPress-1>', lambda e, tlnum=i: edit_task(task=tlnum))

    # add to task popup menu
    for pn, per in enumerate(person):
        per[6].delete(0, 'end')
        for tn, t in enumerate(tasksvariable):
            per[6].add_command(label=t[1],
                               command=lambda row=pn, tasknumber=tn: set_default_task(tasknumber=tasknumber, row=row))

    # re-populate default task combobox
    for employeeswidget in employeeswidgets:
        employeeswidget[3][2] = []
        for t in tasksvariable:
            if t[1]:
                employeeswidget[3][2].append(t[1])
        employeeswidget[3][0]['values'] = employeeswidget[3][2]

    # hide all widgets
    for newtaskwidgetn, newtaskwidget in enumerate(newtaskwidgets):
        if newtaskwidgetn == 15 or newtaskwidgetn == 16:
            for newtaskwidge in newtaskwidget:
                newtaskwidge.grid_remove()
        else:
            newtaskwidget.grid_remove()


def edit_task(task):
    # show all widgets for edit task

    for newtaskwidgetn, newtaskwidget in enumerate(newtaskwidgets):
        if newtaskwidgetn == 15:
            newtaskwidget[task].grid()
        elif newtaskwidgetn == 16:
            if task > 1:
                newtaskwidget[task - 2].grid()
        else:
            newtaskwidget.grid()

    # Put the current value into the widgets
    newtaskwidgets[0]['text'] = tasksvariable[task][1]
    newtaskname.set(tasksvariable[task][1])
    newtaskcolor.set(tasksvariable[task][2])
    newtaskwidgets[4]['bg'] = f'#{newtaskcolor.get()}'
    newtaskauto_generate.set(tasksvariable[task][3])
    newtaskdefault_certified.set(tasksvariable[task][4])
    newtaskschedule_length.set(tasksvariable[task][5])
    newtaskschedule_max_times.set(tasksvariable[task][6])

    # hide new widget save button
    newtaskwidgets[13].grid_remove()


def edit_task_save(task):
    # This function saves the new task to XML, and updates alla associated variables

    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement
    tasks = settings.getElementsByTagName('task')

    # name
    oldname = tasksvariable[task][1]
    tasksvariable[task][1] = newtaskname.get()
    tasklabel[task]['text'] = newtaskname.get()
    taskselector[task]['text'] = newtaskname.get()
    for pn, per in enumerate(person):
        per[6].delete(0, 'end')
        for tn, t in enumerate(tasksvariable):
            if t[1]:
                per[6].add_command(label=t[1],
                                   command=lambda row=pn, tasknumber=tn: set_default_task(tasknumber=tasknumber, row=row))
    tasks[task].getElementsByTagName('name')[0].childNodes[0].nodeValue = tasksvariable[task][1]

    # color
    tasksvariable[task][2] = newtaskcolor.get()
    taskbutton[task]['bg'] = f'#{newtaskcolor.get()}'
    taskbutton[task]['text'] = newtaskcolor.get()
    for p in person:
        for q in p[4]:
            if int(q[1]) == int(tasksvariable[task][0]):
                q[0]['bg'] = f'#{tasksvariable[task][2]}'
        if p[5][1] == tasksvariable[task][0]:
            p[5][0]['bg'] = f'#{tasksvariable[task][2]}'
    tasks[task].getElementsByTagName('color')[0].childNodes[0].nodeValue = tasksvariable[task][2]

    # autogenerate
    tasksvariable[task][3] = newtaskauto_generate.get()
    tasks[task].getElementsByTagName('auto_generate')[0].childNodes[0].nodeValue = str(tasksvariable[task][3])

    # default cerified
    tasksvariable[task][4] = newtaskdefault_certified.get()
    tasks[task].getElementsByTagName('default_certified')[0].childNodes[0].nodeValue = str(tasksvariable[task][4])

    # schedule length
    tasksvariable[task][5] = newtaskschedule_length.get()
    tasks[task].getElementsByTagName('schedule_length')[0].childNodes[0].nodeValue = str(tasksvariable[task][5])

    # schedule max times
    tasksvariable[task][6] = newtaskschedule_max_times.get()
    tasks[task].getElementsByTagName('schedule_max_times')[0].childNodes[0].nodeValue = str(tasksvariable[task][6])

    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")

    # re-populate default task combobox
    for employeeswidget in employeeswidgets:
        employeeswidget[3][2] = []
        for t in tasksvariable:
            if t[1]:
                employeeswidget[3][2].append(t[1])
        employeeswidget[3][0]['values'] = employeeswidget[3][2]

    # calibrate selected default task
    if oldname == employeeswidget[3][1].get():
        employeeswidget[3][1].set(tasksvariable[task][1])

    # hide all widgets
    for newtaskwidgetn, newtaskwidget in enumerate(newtaskwidgets):
        if newtaskwidgetn == 15 or newtaskwidgetn == 16:
            for newtaskwidge in newtaskwidget:
                newtaskwidge.grid_remove()
        else:
            newtaskwidget.grid_remove()


def task_delete(task):
    # Deletes the task from XML, and associated variables

    # check if task is in use
    task_in_use = False
    for per in person:
        for t in per[4]:
            if t[1] == task:
                task_in_use = True

    # Shows error if it's in use
    if task_in_use:
        messagebox.showerror(message='Den här uppgiften används i dagens schema,\noch kan inte tas bort.')

    # Runs if not in use
    else:

        # xml
        xml_delete_task(tasksvariable, task, employees)

        # taskvariable
        tasksvariable[task][1] = ''

        # taskselector
        taskselector[task].grid_remove()

        # task popup
        for pn, per in enumerate(person):
            per[6].delete(0, 'end')
            for tn, t in enumerate(tasksvariable):
                if t[1]:
                    per[6].add_command(label=t[1],
                                       command=lambda row=pn, tasknumber=tn: set_default_task(tasknumber=tasknumber,
                                                                                          row=row))
        # task settings list
        tasklabel[task].grid_remove()
        taskbutton[task].grid_remove()

        # employee settings widgets
        for e in employeeswidgets:
            e[0][0].grid_remove()
            e[0][1].grid_remove()
            e[0][3].grid_remove()
            for w in e[2]:
                w.grid_remove()

        # default task
        for employeeswidget in employeeswidgets:
            employeeswidget[3][2] = []
            for t in tasksvariable:
                if t[1]:
                    employeeswidget[3][2].append(t[1])
            employeeswidget[3][0]['values'] = employeeswidget[3][2]


def settings():
    # The settings window
    global settingsWindow
    global taskbutton
    global tasksframe
    global tasklabel
    taskbutton = []
    tasklabel = []
    settingsWindow = Toplevel(root)
    settingsWindow.title('Settings')
    settingsWindow.attributes("-topmost", 1)
    settingsWindow.resizable(FALSE, FALSE)
    settingsWindow.overrideredirect(True)
    settingsWindow.geometry(f'+580+130')
    title_bar = Frame(settingsWindow, bg='#7070FF', relief='raised')
    title_bar.pack(expand=1, fill=X)
    title_bar.bind("<B1-Motion>", move_settings_window)
    title_label = Label(title_bar, text='Settings', bg='#7070FF', fg='black')
    title_label.pack(side=LEFT, pady=2, padx=2)

    s = ttk.Style()
    s.configure('TNotebook', tabposition='wn')
    tabs = ttk.Notebook(settingsWindow)
    tabs.pack()
    tasksframe = Frame(tabs, relief='ridge', bd=2)
    tasksframe.pack(expand=1, fill=X)
    breakframe = Frame(tabs, relief='ridge', bd=2)
    breakframe.pack(expand=1, fill=X)
    personelframe = Frame(tabs, relief='ridge', bd=2)
    personelframe.pack(expand=1, fill=X)
    excelframe = Frame(tabs, relief='ridge', bd=2)
    excelframe.pack(expand=1, fill=X)
    tabs.add(tasksframe, text='Uppgifter')
    tabs.add(breakframe, text='Rast')
    tabs.add(personelframe, text='Personal')
    tabs.add(excelframe, text='Excel')

    # Tasks tab
    Button(tasksframe, text='Ny uppgift', command=show_new_task).grid(row=998, column=0, columnspan=2, pady=10)
    Button(tasksframe, text='ok', command=settingsWindow.destroy).grid(row=1000, column=1, pady=5)
    taskeditsavebuttons = []
    taskeditdeletebuttons = []
    for i, task in enumerate(tasksvariable):
        tasklabel.append(Label(tasksframe, text=task[1]))
        tasklabel[i].grid(row=i, column=0)
        taskbutton.append(Button(tasksframe,
                                 text=task[2],
                                 bg=f'#{task[2]}'))
        taskbutton[i].grid(row=i, column=1, padx=2, pady=2)
        taskeditsavebuttons.append(Button(tasksframe, text='spara', command=lambda task=i: edit_task_save(task=task)))

        # add a delete button on all tasks except the first 2
        if i > 1:
            taskeditdeletebuttons.append(Button(tasksframe, text='ta bort', command=lambda task=i: task_delete(task=task)))

    for tlnum, tl in enumerate(tasklabel):
        tl.bind('<ButtonPress-1>', lambda e, tlnum=tlnum: edit_task(task=tlnum))
    ttk.Separator(tasksframe, orient=VERTICAL).grid(row=0, rowspan=999, column=2, sticky='wns', padx=2)
    ttk.Separator(tasksframe, orient=HORIZONTAL).grid(row=999, columnspan=5, column=0, sticky='we')
    global newtaskwidgets
    newtaskwidgets = []
    newtaskwidgets.append(Label(tasksframe, text='Lägg till en ny uppgift'))
    newtaskwidgets[0].grid(row=0, column=3, columnspan=2)
    global newtaskname
    newtaskname = StringVar()
    newtaskwidgets.append(Label(tasksframe, text='namn'))
    newtaskwidgets[1].grid(row=1, column=3, sticky='e', padx=2)
    newtaskwidgets.append(Entry(tasksframe, textvariable=newtaskname))
    newtaskwidgets[2].grid(row=1, column=4, sticky='w', padx=2)
    newtaskwidgets.append(Label(tasksframe, text='Färg'))
    newtaskwidgets[3].grid(row=2, column=3, sticky='e', padx=2)
    global newtaskcolor
    newtaskcolor = StringVar()
    newtaskcolor.set('555555')
    newtaskwidgets.append(Button(tasksframe, text='     ', bg=f'#{newtaskcolor.get()}', command=select_color_new_task))
    newtaskwidgets[4].grid(row=2, column=4, sticky='w', padx=2)
    global newtaskauto_generate
    newtaskauto_generate = BooleanVar()
    newtaskauto_generate.set(False)
    newtaskwidgets.append(Label(tasksframe, text='Autogenerera'))
    newtaskwidgets[5].grid(row=3, column=3, sticky='e', padx=2)
    newtaskwidgets.append(ttk.Checkbutton(tasksframe, variable=newtaskauto_generate, onvalue=True, offvalue=False))
    newtaskwidgets[6].grid(row=3, column=4, sticky='w', padx=2)
    global newtaskdefault_certified
    newtaskdefault_certified = BooleanVar()
    newtaskdefault_certified.set(False)
    newtaskwidgets.append(Label(tasksframe, text='Standard behörig'))
    newtaskwidgets[7].grid(row=4, column=3, sticky='e', padx=2)
    newtaskwidgets.append(ttk.Checkbutton(tasksframe, variable=newtaskdefault_certified, onvalue=True, offvalue=False))
    newtaskwidgets[8].grid(row=4, column=4, sticky='w', padx=2)
    global newtaskschedule_length
    newtaskschedule_length = StringVar()
    newtaskschedule_length.set('60')
    newtaskwidgets.append(Label(tasksframe, text='Schemalägg tid (min)'))
    newtaskwidgets[9].grid(row=5, column=3, sticky='e', padx=2)
    newtaskwidgets.append(Entry(tasksframe, textvariable=newtaskschedule_length))
    newtaskwidgets[10].grid(row=5, column=4, sticky='w', padx=2)
    global newtaskschedule_max_times
    newtaskschedule_max_times = StringVar()
    newtaskschedule_max_times.set('3')
    newtaskwidgets.append(Label(tasksframe, text='Schemalägg max antal'))
    newtaskwidgets[11].grid(row=6, column=3, sticky='e', padx=2)
    newtaskwidgets.append(Entry(tasksframe, textvariable=newtaskschedule_max_times))
    newtaskwidgets[12].grid(row=6, column=4, sticky='w', padx=2)
    newtaskwidgets.append(Button(tasksframe, text='spara', command=new_task_save))
    newtaskwidgets[13].grid(row=7, column=3, sticky='e', padx=2)
    newtaskwidgets.append(Button(tasksframe, text='avbryt', command=hide_new_task))
    newtaskwidgets[14].grid(row=7, column=4, sticky='w', padx=2)
    newtaskwidgets.append(taskeditsavebuttons)
    newtaskwidgets.append(taskeditdeletebuttons)
    for newtaskwidget in newtaskwidgets[15]:
        newtaskwidget.grid(row=7, column=3, sticky='e', padx=2)
    for newtaskwidget in newtaskwidgets[16]:
        newtaskwidget.grid(row=7, column=5, sticky='e', padx=2)
    for newtaskwidgetn, newtaskwidget in enumerate(newtaskwidgets):
        if newtaskwidgetn == 15 or newtaskwidgetn == 16:
            for newtaskwidge in newtaskwidget:
                newtaskwidge.grid_remove()
        else:
            newtaskwidget.grid_remove()

    # break tab
    Label(breakframe, text='Arbetad tid i minuter innan rast').grid(column=0, row=0, columnspan=2, padx=4)
    for b in breaksvariable:
        global minbreak
        minbreak = StringVar()
        Label(breakframe, text='Minimum').grid(column=0, row=1)
        Entry(breakframe, textvariable=minbreak, width=6).grid(column=1, row=1)
        minbreak.set(b[0])
        global maxbreak
        maxbreak = StringVar()
        maxbreak.set(b[1])
        Label(breakframe, text='Maximalt').grid(column=0, row=2)
        Entry(breakframe, textvariable=maxbreak, width=6).grid(column=1, row=2)
        ttk.Separator(breakframe, orient=HORIZONTAL).grid(row=3, columnspan=2, column=0, sticky='wes', pady=2)

        breakslengthframe = Frame(breakframe)
        breakslengthframe.grid(column=0, row=4, columnspan=2, rowspan=12)
        Label(breakslengthframe, text='Rastlängder').grid(column=0, row=4, columnspan=5, padx=4)
        i = 0
        global breakslengthsettings
        breakslengthsettings = []
        for bl in breakslength:
            temp = [StringVar(), StringVar(), StringVar(), StringVar()]
            Label(breakslengthframe, text=f'{i + 4} timmar').grid(column=0, row=i + 5)
            temp[0].set(bl[0])
            temp[1].set(bl[1])
            temp[2].set(bl[2])
            temp[3].set(bl[3])
            breakslengthsettings.append(temp)
            Entry(breakslengthframe, textvariable=breakslengthsettings[i][0], width=3).grid(column=1, row=i + 5)
            Entry(breakslengthframe, textvariable=breakslengthsettings[i][1], width=3).grid(column=2, row=i + 5)
            Entry(breakslengthframe, textvariable=breakslengthsettings[i][2], width=3).grid(column=3, row=i + 5)
            Entry(breakslengthframe, textvariable=breakslengthsettings[i][3], width=3).grid(column=4, row=i + 5)
            i = i + 1

    Button(breakframe, text='spara', command=save_break).grid(row=1000, column=0, pady=4)
    Button(breakframe, text='ok', command=settingsWindow.destroy).grid(row=1000, column=1, pady=4)

    ttk.Separator(breakframe, orient=VERTICAL).grid(row=0, rowspan=999, column=2, sticky='wns')
    ttk.Separator(breakframe, orient=HORIZONTAL).grid(row=999, columnspan=5, column=0, sticky='wnse', padx=4)

    Label(breakframe, text='Minsta antalet personal').grid(column=3, row=0, columnspan=2, padx=4)
    i = 0
    global worker
    worker = []
    for w in workersminimum:
        Label(breakframe, text=f'{str(i + 8)}-{str(i + 9)}').grid(column=3, row=i + 1, sticky='ens', padx=2)
        worker.append(StringVar())
        worker[i].set(w)
        Entry(breakframe, textvariable=worker[i], width=2).grid(column=4, row=i + 1, sticky='wns', padx=2)
        i = i + 1

    # peronelframe
    Button(personelframe, text='ok', command=settingsWindow.destroy).grid(row=1001, column=0, pady=4)
    global employeeswidgets
    employeeswidgets = []
    global employeeslistboxitems
    employeeslistboxitems = []
    ttk.Separator(personelframe, orient=VERTICAL).grid(row=0, rowspan=1001, column=2, sticky='wns', padx=5)
    for i, employee in enumerate(employees):
        tempvar = []
        tempwidgets = []
        employeeslistboxitems.append(employee[0])
        tasknumber = 0
        for j, t in enumerate(employee[1]):
            for k, task in enumerate(tasksvariable):
                if t[0] == task[0]:
                    tasknumber = k
            tempvar.append(BooleanVar())
            tempvar[j].set(t[1])
            tempwidgets.append(ttk.Checkbutton(personelframe,
                                               text=tasksvariable[tasknumber][1],
                                               variable=tempvar[j],
                                               onvalue=True,
                                               offvalue=False))
            tempwidgets[j].grid(column=3, row=1 + j, sticky='w')
            tempwidgets[j].grid_remove()

        # populate default task combobox and set current value
        default_task_variable = StringVar()
        default_task_options = []
        for dtask in tasksvariable:
            default_task_options.append(dtask[1])
            if dtask[0] == employee[2]:
                default_task_variable.set(dtask[1])

        employeeswidgets.append([[Label(personelframe, text=employee[0]),
                                  ttk.Button(personelframe, text='spara', command=lambda row=i: save_employee(row=row)),
                                  employee[0], ttk.Button(personelframe,
                                                          text='ta bort',
                                                          command=lambda row=i: delete_employee(row=row))],
                                 tempvar,
                                 tempwidgets,
                                 [ttk.Combobox(personelframe, textvariable=default_task_variable, state='readonly'), default_task_variable, default_task_options],
                                 Label(personelframe, text='Standarduppgift')])
        employeeswidgets[i][0][0].grid(column=3, row=0)
        employeeswidgets[i][0][0].grid_remove()
        employeeswidgets[i][0][1].grid(column=3, row=1000)
        employeeswidgets[i][0][1].grid_remove()
        employeeswidgets[i][0][3].grid(column=4, row=1000)
        employeeswidgets[i][0][3].grid_remove()

        # default task combo box
        employeeswidgets[i][3][0]['values'] = default_task_options
        employeeswidgets[i][3][0].grid(column=3, row=998)
        employeeswidgets[i][3][0].grid_remove()
        employeeswidgets[i][4].grid(column=3, row=997)
        employeeswidgets[i][4].grid_remove()

    global employeeslistboxitemsvar
    employeeslistboxitemsvar = StringVar(value=employeeslistboxitems)
    global employeeslistbox
    employeeslistbox = Listbox(personelframe, listvariable=employeeslistboxitemsvar, height=18, exportselection=False)
    employeeslistbox.grid(column=0, row=0, rowspan=20)
    employeeslistbox.bind('<<ListboxSelect>>', show_employee)

    # excelframe
    global excellwidgets
    excellwidgets = []
    var = StringVar()
    var.set(excel_selected_variable[0])
    excellwidgets.append(var)
    tempwidgets = []
    for excell_key in excell_templates:
        tempwidgets.append(ttk.Radiobutton(excelframe, variable=excellwidgets[0], text=excell_templates[excell_key][0], value=excell_key))
    for wn, w in enumerate(tempwidgets):
        w.grid(column=0, row=wn, sticky='w')
    excellwidgets.append(tempwidgets)
    Button(excelframe, text='lägg till ny mall', command=show_add_excel).grid(row=999, column=0, pady=4)
    Button(excelframe, text='spara', command=save_excel).grid(row=1000, column=0, pady=4)
    Button(excelframe, text='ok', command=settingsWindow.destroy).grid(row=1000, column=1, pady=4)
    ttk.Separator(excelframe, orient=VERTICAL).grid(row=0, rowspan=1001, column=2, sticky='wns', padx=5)

    # widgets on add_excel
    global add_excel_widgets
    add_excel_widgets = []
    global add_excel_variables
    add_excel_variables = []

    add_excel_widgets.append(Label(excelframe, text='Namn: '))
    add_excel_widgets[0].grid(column=3, row=0)

    add_excel_variables.append(StringVar())
    add_excel_variables[0].set('Ny mall')
    add_excel_widgets.append(Entry(excelframe, textvariable=add_excel_variables[0]))
    add_excel_widgets[1].grid(column=4, row=0)

    add_excel_widgets.append(Button(excelframe, text='Öppna fil', command=select_add_excel_file))
    add_excel_widgets[2].grid(column=3, row=1)

    add_excel_widgets.append(Label(excelframe, text=' '))
    add_excel_widgets[3].grid(column=3, row=2)

    add_excel_widgets.append(Button(excelframe, text='Lägg till', command=save_add_excel))
    add_excel_widgets[4].grid(column=3, row=3)

    for add_excel_widget in add_excel_widgets:
        add_excel_widget.grid_remove()


def save_add_excel():
    # Adds a new xlsx document to the xml, and updates the associated variables

    # Load the excel-spreadsheet
    wb = load_workbook(excel_add_file_name)
    ws = wb.active

    # add xlsx to xml
    data, excel_id = xml_save_excel_template(ws, excell_templates, add_excel_variables)

    # update excel-variables
    excell_templates[str(excel_id)] = [add_excel_variables[0].get(), data]


def select_add_excel_file():
    global excel_add_file_name
    excel_add_file_name = filedialog.askopenfilename(initialdir='.', filetypes=(('excel','*.xlsx'), ('all files', '*.*')))
    file_name = excel_add_file_name.split("/")
    add_excel_widgets[3]['text'] = file_name[len(file_name) - 1]


def show_add_excel():
    for add_excel_widget in add_excel_widgets:
        add_excel_widget.grid()


def save_excel():
    # saves selected excel template

    xml_save_excel(excellwidgets)
    excel_selected_variable[0] = excellwidgets[0].get()


def delete_employee(row):
    # create an array of all active employees
    activeemployees = []
    for per in person:
        if per[0].get():
            activeemployees.append(per[5])

    # check if employee is active on todays schedule
    if row in activeemployees:
        messagebox.showerror(message='Den här personen finns med på dagens schema, och kan inte tas bort')

    # if employee is not on todays schedule
    else:

        # load xml
        domtree = xml.dom.minidom.parse('settings.xml')
        settings = domtree.documentElement
        employee = settings.getElementsByTagName('employee')
        for e in employee:
            name = e.getElementsByTagName('name')[0].childNodes[0].nodeValue

            # remove from xml if the selected employee is found
            if name == employees[row][0]:
                e.parentNode.removeChild(e)

        # save xml with utf-8
        domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")

        # Clear the name of the employee to deactivate him/her
        employees[row][0] = ''

        # re-populate employeeslistboxitems
        employeeslistboxitems.clear()
        for employee in employees:
            if employee[0] != '':
                employeeslistboxitems.append(employee[0])
        employeeslistboxitemsvar.set(value=employeeslistboxitems)


def show_employee(*args):
    for e in employeeswidgets:
        e[0][0].grid_remove()
        e[0][1].grid_remove()
        e[0][3].grid_remove()
        e[3][0].grid_remove()
        e[4].grid_remove()
        for w in e[2]:
            w.grid_remove()
    emp = employeeslistbox.curselection()[0]
    employeeswidgets[emp][0][0].grid()
    employeeswidgets[emp][0][1].grid()
    employeeswidgets[emp][0][3].grid()
    employeeswidgets[emp][3][0].grid()
    employeeswidgets[emp][4].grid()
    for t, w in enumerate(employeeswidgets[emp][2]):
        if tasksvariable[t][1]:
            w.grid()
    return True


def save_employee(row):
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement
    emps = settings.getElementsByTagName('employee')
    for emp in emps:
        name = emp.getElementsByTagName('name')
        if name[0].childNodes[0].nodeValue == employeeswidgets[row][0][2]:

            # save default task
            for task in tasksvariable:
                if task[1] == employeeswidgets[row][3][1].get():
                    emp.getElementsByTagName('default_task')[0].childNodes[0].nodeValue = task[0]

            ts = emp.getElementsByTagName('task_settings')
            for i, t in enumerate(ts):
                t.getElementsByTagName('certified')[0].childNodes[0].nodeValue = employeeswidgets[row][1][i].get()
                employees[row][1][i][1] = str(employeeswidgets[row][1][i].get())

    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")


def save_break():
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement
    breaks = settings.getElementsByTagName('break')
    breaks[0].getElementsByTagName('min')[0].childNodes[0].nodeValue = minbreak.get()
    breaks[0].getElementsByTagName('max')[0].childNodes[0].nodeValue = maxbreak.get()
    breaksvariable[0][0] = minbreak.get()
    breaksvariable[0][1] = maxbreak.get()

    workers = settings.getElementsByTagName('workers_minimum')
    for i in range(13):
        workers[0].getElementsByTagName(f'h{i + 8}')[0].childNodes[0].nodeValue = worker[i].get()
        workersminimum[i] = worker[i].get()

    workingtimes = settings.getElementsByTagName('workingtime')
    for i in range(6):
        breaktimes = workingtimes[0].getElementsByTagName(f'h{i + 4}')
        breaktimes[0].getElementsByTagName('first_break')[0].childNodes[0].nodeValue = breakslengthsettings[i][0].get()
        breaktimes[0].getElementsByTagName('second_break')[0].childNodes[0].nodeValue = breakslengthsettings[i][1].get()
        breaktimes[0].getElementsByTagName('third_break')[0].childNodes[0].nodeValue = breakslengthsettings[i][2].get()
        breaktimes[0].getElementsByTagName('forth_break')[0].childNodes[0].nodeValue = breakslengthsettings[i][3].get()
        breakslength[i][0] = breakslengthsettings[i][0].get()
        breakslength[i][1] = breakslengthsettings[i][1].get()
        breakslength[i][2] = breakslengthsettings[i][2].get()
        breakslength[i][3] = breakslengthsettings[i][3].get()

    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")


def export_to_excel():


    # logging
    if log['export_to_excel']:
        time = datetime.datetime.now()
        excel_template_title = excell_templates[excel_selected_variable[0]][0]
        with open(logfile, 'a') as f:
            f.write(f'{time.hour}:{time.minute}:{time.second} export_to_excel: template: {excel_template_title}\n')

    wb = Workbook()
    ws = wb.active
    ws.title = "Schema"
    sidethin = Side(border_style='thin')
    sidethick = Side(border_style='medium')
    borderboxthin = Border(top=sidethin, bottom=sidethin, left=sidethin, right=sidethin)
    borderboxthick = Border(top=sidethick, bottom=sidethick, left=sidethick, right=sidethick)
    bordertopbottomthick = Border(top=sidethick, bottom=sidethick)
    bordertopbottomrightthick = Border(top=sidethick, bottom=sidethick, right=sidethick)
    borderrightthicktopbottomthin = Border(right=sidethick, top=sidethin, bottom=sidethin)
    bordertopthick = Border(top=sidethick)
    ws.column_dimensions[get_column_letter(1)].width = 12.4
    ws.column_dimensions[get_column_letter(2)].width = 11.8
    for i in range(52):
        ws.column_dimensions[get_column_letter(i + 3)].width = 2.8

    # setup page to landscape A4 and to fit a single page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.page_setup.paperSize = 9

    ws['A1'] = 'Namn'
    ws['A1'].border = borderboxthick
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 'Arbetstid'
    ws['B1'].border = borderboxthick
    ws['B1'].font = Font(bold=True)
    ws['C1'] = '08:00-09:00'
    ws['G1'] = '09:00-10:00'
    ws['K1'] = '10:00-11:00'
    ws['O1'] = '11:00-12:00'
    ws['S1'] = '12:00-13:00'
    ws['W1'] = '13:00-14:00'
    ws['AA1'] = '14:00-15:00'
    ws['AE1'] = '15:00-16:00'
    ws['AI1'] = '16:00-17:00'
    ws['AM1'] = '17:00-18:00'
    ws['AQ1'] = '18:00-19:00'
    ws['AU1'] = '19:00-20:00'
    ws['AY1'] = '20:00-21:00'
    for col in range(52):
        ws[f'{get_column_letter(col + 3)}1'].border = bordertopbottomthick
    for col in range(13):
        ws[f'{get_column_letter((col + 1) * 4 + 2)}1'].border = bordertopbottomrightthick
    i = 1
    for per in person:
        if per[0].get():
            ws[f'A{i + 1}'] = per[0].get().lower().capitalize()
            ws[f'B{i + 1}'] = per[2].get()
            ws[f'A{i + 1}'].border = borderboxthick
            ws[f'B{i + 1}'].border = borderboxthick
            j = 3
            for status in per[4]:
                color = 'FFFFFF'
                if status[1] > -1:
                    color = f'FF{tasksvariable[status[1]][2]}'
                col = get_column_letter(j)
                ws[col + str(i + 1)].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws[col + str(i + 1)].border = borderboxthin
                ws[col + str(i + 1)] = status[0]['text']
                ws[col + str(i + 1)].alignment = Alignment(horizontal='center')
                ws[col + str(i + 1)].font = Font(bold=True)

                j = j + 1

            for col in range(13):
                ws[f'{get_column_letter((col + 1) * 4 + 2)}{i + 1}'].border = borderrightthicktopbottomthin

            for col in range(54):
                ws[f'{get_column_letter(col + 1)}{i + 2}'].border = bordertopthick
            i = i + 1
    tasknumber = 0
    activetasks = []
    for row in range(len(person)):
        for col in range(52):
            if not person[row][4][col][1] + 1 in activetasks:
                activetasks.append(person[row][4][col][1] + 1)
    for task in tasksvariable:
        if int(task[0]) in activetasks:
            ws[f'{get_column_letter(tasknumber * 6 + 3)}{i + 2}'].fill = PatternFill(start_color=f'FF{task[2]}',
                                                                                     end_color=f'FF{task[2]}',
                                                                                     fill_type="solid")
            ws[f'{get_column_letter(tasknumber * 6 + 4)}{i + 2}'] = task[1]
            tasknumber = tasknumber + 1

    # add footer
    row = i + 3  # calculate starting row for footer
    for cell in excell_templates[excel_selected_variable[0]][1]:
        cell_adress = cell['id'].split(':')

        # text
        ws[f'{cell_adress[0]}{int(cell_adress[1]) + row}'] = cell['text']

        # background color
        if cell['bg']:
            color = cell['bg']
        else:
            color = 'FFFFFFFF'
        ws[f'{cell_adress[0]}{int(cell_adress[1]) + row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        #border
        if cell['border_left']:
            if cell['border_left_color']:
                color = f"00{cell['border_left_color']}"
            else:
                color = '00000000'
            border_left = Side(border_style=cell['border_left'], color=color)
        else:
            border_left = Side(border_style='none')
        if cell['border_right']:
            if cell['border_right_color']:
                color = f"00{cell['border_right_color']}"
            else:
                color = '00000000'
            border_right = Side(border_style=cell['border_right'], color=color)
        else:
            border_right = Side(border_style='none')
        if cell['border_top']:
            if cell['border_top_color']:
                color = f"00{cell['border_top_color']}"
            else:
                color = '00000000'
            border_top = Side(border_style=cell['border_top'], color=color)
        else:
            border_top = Side(border_style='none')
        if cell['border_bottom']:
            if cell['border_bottom_color']:
                color = f"00{cell['border_bottom_color']}"
            else:
                color = '00000000'
            border_bottom = Side(border_style=cell['border_bottom'], color=color)
        else:
            border_bottom = Side(border_style='none')
        ws[f'{cell_adress[0]}{int(cell_adress[1]) + row}'].border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom)

        # font, font size, font style, fg
        if cell['font_style_bold']:
            font_style_bold = eval(cell['font_style_bold'])
        else:
            font_style_bold = False
        if cell['font_style_italic']:
            font_style_italic = cell['font_style_italic']
        else:
            font_style_italic = False
        if cell['font_style_underline']:
            font_style_underline = cell['font_style_underline']
        else:
            font_style_underline = None
        if cell['fg']:
            fg = cell['fg']
        else:
            fg = '000000'
        if cell['font_size']:
            font_size = cell['font_size']
        else:
            font_size = None
        if cell['font']:
            font = cell['font']
        else:
            font = None
        ws[f'{cell_adress[0]}{int(cell_adress[1]) + row}'].font = Font(bold=font_style_bold,
                                                                       italic=font_style_italic,
                                                                       underline=font_style_underline,
                                                                       color=fg,
                                                                       size=font_size,
                                                                       name=font)

    wb.save('schema.xlsx')


def show_task_popup(e, row):
    person[row][6].tk_popup(e.x_root, e.y_root)


def set_default_task(tasknumber, row):
    # logging
    if log['set_default_task']:
        time = datetime.datetime.now()
        task = tasksvariable[tasknumber][1]
        with open(logfile, 'a') as f:
            f.write(f'{time.hour}:{time.minute}:{time.second} set_default_task: row: {row} task: {task}\n')

    for col in person[row][4]:
        if int(col[1]) > -1:
            col[0]['bg'] = f'#{tasksvariable[tasknumber][2]}'
            col[1] = tasknumber
    person[row][5][0]['bg'] = f'#{tasksvariable[tasknumber][2]}'
    person[row][5][1] = tasknumber


def about():
    about_window = Toplevel(root)
    about_window.title('Om')
    about_window.resizable(FALSE, FALSE)
    about_window.geometry('200x150')
    Label(about_window, text=f'Rastplaneraren v. {version}').grid(column=1, row=5, padx=10, pady=10)


def show_announcements(announcements):

    # logging
    if log['show_announcements']:
        time = datetime.datetime.now()
        with open(logfile, 'a') as f:
            f.write(f'{time.hour}:{time.minute}:{time.second} show_announcements\n')

    announcements_window = Toplevel(root)
    announcements_window.title('Nyheter')
    announcements_window.attributes("-topmost", 1)
    announcements_variables = []

    for i, a in enumerate(announcements):
        announcements_variables.append(BooleanVar())
        Checkbutton(announcements_window, text='läst',
                    variable=announcements_variables[i], offvalue=False, onvalue=True).grid(row=i, column=0)
        Label(announcements_window, text=a, justify='left').grid(row=i, column=3, padx=4)
    Button(announcements_window, text='ok',
           command=lambda announcements_variables=announcements_variables, announcements_window=announcements_window: hide_announcements(announcements_variables, announcements_window))\
        .grid(row=i+1, column=0, columnspan=4, pady=5)
    

def hide_announcements(announcements_variables, announcements_window):
    # Closes announcements window and removes selected ones

    # Delete announcement if it's marked as read
    for i, a in enumerate(announcements_variables):
        if a.get():
            # logging
            if log['delete_announcements']:
                time = datetime.datetime.now()
                with open(logfile, 'a') as f:
                    f.write(f'{time.hour}:{time.minute}:{time.second} delete_announcement: {i}\n')

            delete_announcement(i)
    announcements_window.destroy()


def add_row():

    if not person:
        row = 0
    else:
        row = len(person)
    temp = []
    temp.append(StringVar())
    temp.append(ttk.Entry(scrollable_middleframe,
                          textvariable=temp[0],
                          validate="all",
                          validatecommand=(addPerson_wrapper, row, "%P", "%V")))
    temp.append(StringVar())
    temp.append(ttk.Entry(scrollable_middleframe,
                          textvariable=temp[2],
                          validate="all",
                          width=11,
                          validatecommand=(addTime_wrapper, row, "%P", "%V")))
    button_inner = []
    for j in range(52):
        button_inner.append([Button(scrollable_middleframe,
                                    height=1,
                                    width=2,
                                    bg='#54FA9B',
                                    command=lambda row=row, col=j: button_color(row=row, col=col)), -1])
    temp.append(button_inner)
    temp.append([Label(scrollable_middleframe, text='->', bg=f'#{tasksvariable[0][2]}'), 0])
    temp.append(Menu(scrollable_middleframe, tearoff=False))
    person.append(temp)
    scrollable_middleframe.rowconfigure(row + 1, minsize=28)
    person[row][1].grid(column=0, row=row + 1, sticky='wns', pady=1, padx=1)
    person[row][3].grid(column=1, row=row + 1, sticky='wns', pady=1, padx=1)
    person[row][5][0].grid(column=2, row=row + 1, sticky='wns', pady=1, padx=2)
    for j in range(52):
        person[row][4][j][0].grid(column=j + 3, row=row + 1, sticky='wn', padx=1, pady=1)
        person[row][4][j][0].grid_remove()
    for tasknumber, availabletask in enumerate(tasksvariable):
        person[row][6].add_command(label=availabletask[1],
                                 command=lambda row=row, tasknumber=tasknumber: set_default_task(tasknumber=tasknumber,
                                                                                               row=row))
    person[row][5][0].bind('<Button-1>', lambda e, row=row: show_task_popup(e=e, row=row))



tasksvariable = []
breaksvariable = []
workersminimum = []
breakslength = []
employees = []
taskselector = []
announcements = []
excell_templates = {}
excel_selected_variable = ['0']
getsettings(tasksvariable, breaksvariable, workersminimum, breakslength, employees, version, excell_templates, excel_selected_variable, announcements)
employees = sorted(employees)
root = Tk()
addTime_wrapper = root.register(add_time)
addPerson_wrapper = root.register(add_person)
root.geometry("1650x600")
person = []

root.title("Rastplaneraren")
menubar = Menu(root)
arkivmenu = Menu(menubar, tearoff=0)
arkivmenu.add_command(label="Avsluta", command=root.quit)
menubar.add_cascade(label="Arkiv", menu=arkivmenu)
verktygmenu = Menu(menubar, tearoff=0)
verktygmenu.add_command(label='Inställningar', command=settings)
menubar.add_cascade(label='Verktyg', menu=verktygmenu)
helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label='Om', command=about)
menubar.add_cascade(label='Hjälp', menu=helpmenu)
root.config(menu=menubar)

if len(announcements):
    show_announcements(announcements)

separatorStyle = ttk.Style()
separatorStyle.configure('TSeparator', background='black')

topframe = ttk.Frame(root, padding="3 3 3 3")
topframe.grid(column=0, row=0, sticky='nwes')
activetask = IntVar()

for i, task in enumerate(tasksvariable):
    taskselector.append(ttk.Radiobutton(topframe, text=task[1], variable=activetask, value=i))
    taskselector[i].grid(column=i, row=0)

activetask.set(1)

middleframe = ttk.Frame(root, padding="3 3 3 3")
middleframe.grid(column=0, row=1, sticky='nwes')

# scrollbar
canvas = Canvas(middleframe, borderwidth=0, border=0)
middleframe_scrollbar = ttk.Scrollbar(middleframe, orient="vertical", command=canvas.yview)
scrollable_middleframe = ttk.Frame(canvas)
scrollable_middleframe.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion=canvas.bbox("all")
    )
)
canvas.create_window((0, 0), window=scrollable_middleframe, anchor="nw")
canvas.configure(yscrollcommand=middleframe_scrollbar.set)
canvas.grid(row=0, column=0, sticky='nwse')
middleframe_scrollbar.grid(row=0, column=1, sticky='nes')

add_row()

for k in range(14):
    ttk.Separator(scrollable_middleframe,
                  orient=VERTICAL).grid(column=(3 + 4 * k), row=1, rowspan=1000, sticky='wns', padx=0)

# Name and working hours headlines
ttk.Label(scrollable_middleframe, text='Namn').grid(row=0, column=0, sticky='w')
ttk.Label(scrollable_middleframe, text='Arbetstid').grid(row=0, column=1, sticky='w')

# Time headlines
for i in range(14):
    if i == 0:
        ttk.Label(scrollable_middleframe, text=str(i + 8)).grid(row=0, column=i * 4 + 3, sticky='w')
    else:
        ttk.Label(scrollable_middleframe, text=str(i + 8)).grid(row=0, column=i * 4 + 2, columnspan=2)

bottomframe = ttk.Frame(root, padding="3 3 3 3", height=50)
bottomframe.grid(column=0, row=2, sticky='nwes')

Label(bottomframe, text=' ').grid(row=0, column=2)
generateoptions = [BooleanVar(value=True), BooleanVar(value=False), BooleanVar(value=False)]
ttk.Checkbutton(bottomframe, text='rast', variable=generateoptions[0], onvalue=True, offvalue=False).grid(row=1,
                                                                                                          column=0,
                                                                                                          sticky='w')
ttk.Checkbutton(bottomframe, text='uppgift', variable=generateoptions[1], onvalue=True, offvalue=False).grid(row=2,
                                                                                                             column=0,
                                                                                                             sticky='w')
ttk.Checkbutton(bottomframe,
                text='prioritering',
                variable=generateoptions[2],
                onvalue=True, offvalue=False).grid(row=3,
                                                   column=0,
                                                   sticky='w')
ttk.Button(bottomframe,
           text='planera\n raster',
           command=lambda generateoptions=generateoptions,
                          person=person,
                          breakslength=breakslength,
                          breaksvariable=breaksvariable,
                          workersminimum=workersminimum,
                          tasksvariable=tasksvariable,
                          employees=employees: plan_breaks(generateoptions,
                                                           person,
                                                           breakslength,
                                                           breaksvariable,
                                                           workersminimum,
                                                           tasksvariable,
                                                           employees))\
    .grid(row=1, column=1, rowspan=3, ipady=12)

Label(bottomframe, text=' ').grid(row=0, column=2, padx=10)
ttk.Button(bottomframe, text='exportera\n till excell', command=export_to_excel).grid(row=1, column=3, rowspan=3,
                                                                                      ipady=12)

# configure grid size
root.rowconfigure(1, weight=1)
root.columnconfigure(0, weight=1)
for j in range(52):
    scrollable_middleframe.columnconfigure(j + 2, minsize=26)
middleframe.columnconfigure(0, weight=1)
middleframe.rowconfigure(0, weight=1)
canvas.columnconfigure(0, weight=1)
canvas.rowconfigure(0, weight=1)

root.mainloop()

# logging
if log['start_stop']:
    time = datetime.datetime.now()
    with open(logfile, 'a') as f:
        f.write(f'{time.hour}:{time.minute}:{time.second} start_stop: program stopped\n')

