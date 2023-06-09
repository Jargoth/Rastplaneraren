# This module contains all functions that handles the XML-file containing the settings

import xml.dom.minidom
import codecs
from openpyxl.utils import get_column_letter
import datetime

import log_system

log, logfile = log_system.start()

# project modules
import default_settings

def update_version(version):
    # Changes for different versions

    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement

    # Load current version
    v = settings.getAttribute('version')
    v = v.split('.')

    # upgrade to version 0.1.3
    if int(v[0]) <= 0 and int(v[1]) <= 1 and int(v[2]) < 3:
        settings.setAttribute('version', version)
        announcement = domtree.createElement('announcement')
        announcement.appendChild(domtree.createTextNode(
            'Nytt i version 0.1.3.\n*Felmeddelande när man försöker generera\n ett schema utan några val.\n*Meddelande om förändringar i nya versioner.\n*Åtgärdat ett fel där minimiarbetstiden\n innan rast inte alltid efterlevdes\n vid rastplanering.'))
        settings.appendChild(announcement)

        domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")

    # upgrade to version 0.1.4
    if int(v[0]) <= 0 and int(v[1]) <= 1 and int(v[2]) < 4:
        settings.setAttribute('version', version)
        announcement = domtree.createElement('announcement')
        announcement.appendChild(domtree.createTextNode('Nytt i version 0.1.4.\n*Nu loggas felen i mappen logs\n*Fixat ett fel i versionsmeddelandena.\n*Loggar fel.\n*Loggar viktiga händelser\n*Validering av datan som skrivs in på namn\n på dagens schema\n*Designförbättring på dagens schema.\n*Prestandaförbättring på dagens schema.\n*Godtyckligt antal personer på dagens schema\nistället för 15.'))
        settings.appendChild(announcement)

        domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")

def getsettings(tasksvariable, breaksvariable, workersminimum, breakslength, employees, version, excell_templates, excel_selected_variable, announcements):
    try:
        domtree = xml.dom.minidom.parse('settings.xml')
    except:
        domtree = default_settings.set_default(version)
    settings = domtree.documentElement

    update_version(version)

    tasks = settings.getElementsByTagName('task')
    for task in tasks:
        temp = []
        temp.append(task.getAttribute('id'))
        temp.append(task.getElementsByTagName('name')[0].childNodes[0].nodeValue)
        temp.append(task.getElementsByTagName('color')[0].childNodes[0].nodeValue)
        temp.append(
            bool(eval(task.getElementsByTagName('auto_generate')[0].childNodes[0].nodeValue.lower().capitalize())))
        temp.append(
            bool(eval(task.getElementsByTagName('default_certified')[0].childNodes[0].nodeValue.lower().capitalize())))
        temp.append(task.getElementsByTagName('schedule_length')[0].childNodes[0].nodeValue)
        temp.append(task.getElementsByTagName('schedule_max_times')[0].childNodes[0].nodeValue)

        # logging
        if log['load_task']:
            time = datetime.datetime.now()
            with open(logfile, 'a') as f:
                f.write(f'{time.hour}:{time.minute}:{time.second} load_task: id: {temp[0]} name: {temp[1]} color: {temp[2]} auto_generate: {temp[3]} default_certified: {temp[4]} schedule_length: {temp[5]} schedule_max_times: {temp[6]}\n')

        tasksvariable.append(temp)

    breaks = settings.getElementsByTagName('break')
    for b in breaks:

        # logging
        if log['load_breaks']:
            time = datetime.datetime.now()
            min = b.getElementsByTagName('min')[0].childNodes[0].nodeValue
            max = b.getElementsByTagName('max')[0].childNodes[0].nodeValue
            with open(logfile, 'a') as f:
                f.write(f'{time.hour}:{time.minute}:{time.second} load_breaks: min: {min} max: {max}\n')

        temp = []
        temp.append(b.getElementsByTagName('min')[0].childNodes[0].nodeValue)
        temp.append(b.getElementsByTagName('max')[0].childNodes[0].nodeValue)
        breaksvariable.append(temp)

    w = settings.getElementsByTagName('workers_minimum')
    for i in range(13):

        # logging
        if log['load_breaks']:
            time = datetime.datetime.now()
            workers = w[0].getElementsByTagName('h' + str(i + 8))[0].childNodes[0].nodeValue
            with open(logfile, 'a') as f:
                f.write(f'{time.hour}:{time.minute}:{time.second} load_workers_min: hour: {i + 8}-{i+9} workers: {workers}\n')

        workersminimum.append(w[0].getElementsByTagName('h' + str(i + 8))[0].childNodes[0].nodeValue)

    workingtime = settings.getElementsByTagName('workingtime')
    for i in range(12):
        c = []
        b = workingtime[0].getElementsByTagName('h' + str(i + 4))

        # logging
        if log['load_working_time']:
            time = datetime.datetime.now()
            first_break = b[0].getElementsByTagName('first_break')[0].childNodes[0].nodeValue
            second_break = b[0].getElementsByTagName('second_break')[0].childNodes[0].nodeValue
            third_break = b[0].getElementsByTagName('third_break')[0].childNodes[0].nodeValue
            forth_break = b[0].getElementsByTagName('forth_break')[0].childNodes[0].nodeValue
            with open(logfile, 'a') as f:
                f.write(f'{time.hour}:{time.minute}:{time.second} load_working_time: working_time: {str(i + 4)} first_break: {first_break} second_break: {second_break} third_break: {third_break} forth_break: {forth_break}\n')

        c.append(b[0].getElementsByTagName('first_break')[0].childNodes[0].nodeValue)
        c.append(b[0].getElementsByTagName('second_break')[0].childNodes[0].nodeValue)
        c.append(b[0].getElementsByTagName('third_break')[0].childNodes[0].nodeValue)
        c.append(b[0].getElementsByTagName('forth_break')[0].childNodes[0].nodeValue)
        breakslength.append(c)

    employee = settings.getElementsByTagName('employee')
    for e in employee:
        name = e.getElementsByTagName('name')[0].childNodes[0].nodeValue
        default_task = e.getElementsByTagName('default_task')[0].childNodes[0].nodeValue
        temp = []
        tasks = e.getElementsByTagName('task_settings')
        temp_string = ''
        for task in tasks:
            temp.append([task.getAttribute('id'), task.getElementsByTagName('certified')[0].childNodes[0].nodeValue])
            temp_string = f"{temp_string} task_id: {task.getAttribute('id')} task_certified: {task.getElementsByTagName('certified')[0].childNodes[0].nodeValue}"
        employees.append([name, temp, default_task])

        # logging
        if log['load_employees']:
            time = datetime.datetime.now()
            with open(logfile, 'a') as f:
                f.write(f'{time.hour}:{time.minute}:{time.second} load_employees: name: {name}{temp_string}\n')

    announcement = settings.getElementsByTagName('announcement')
    for a in announcement:

        # logging
        if log['load_announcements']:
            time = datetime.datetime.now()
            with open(logfile, 'a') as f:
                f.write(f'{time.hour}:{time.minute}:{time.second} load_announcements: {a.childNodes[0].nodeValue}\n')

        announcements.append(a.childNodes[0].nodeValue)

    # load excell template
    excell_templates['0'] = ['Empty', []]
    excells = settings.getElementsByTagName('excell')
    for excell in excells:
        data = []
        cells = excell.getElementsByTagName('cell')
        for cell in cells:
            temp = {}
            temp['id'] = cell.getAttribute('id')
            temp['font'] = cell.getAttribute('font')
            temp['font_size'] = cell.getAttribute('font_size')
            temp['font_style_bold'] = cell.getAttribute('font_style_bold')
            temp['font_style_italic'] = cell.getAttribute('font_style_italic')
            temp['font_style_underline'] = cell.getAttribute('font_style_underline')
            temp['fg'] = cell.getAttribute('fg')
            temp['bg'] = cell.getAttribute('bg')
            temp['border_left'] = cell.getAttribute('border_left')
            temp['border_right'] = cell.getAttribute('border_right')
            temp['border_top'] = cell.getAttribute('border_top')
            temp['border_bottom'] = cell.getAttribute('border_bottom')
            temp['border_left_color'] = cell.getAttribute('border_left_color')
            temp['border_right_color'] = cell.getAttribute('border_right_color')
            temp['border_top_color'] = cell.getAttribute('border_top_color')
            temp['border_bottom_color'] = cell.getAttribute('border_bottom_color')
            if cell.childNodes:
                temp['text'] = cell.childNodes[0].nodeValue
            else:
                temp['text'] = ''
            data.append(temp)

            # logging
            if log['load_excel_templates']:
                time = datetime.datetime.now()
                with open(logfile, 'a') as f:
                    f.write(
                        f'{time.hour}:{time.minute}:{time.second} load_excel_templates: cell_data: {temp}\n')

        title = excell.getElementsByTagName('title')
        excell_templates[excell.getAttribute('id')] = [title[0].childNodes[0].nodeValue, data]

        # logging
        if log['load_excel_templates']:
            time = datetime.datetime.now()
            with open(logfile, 'a') as f:
                f.write(f"{time.hour}:{time.minute}:{time.second} load_excel_templates: title: {title[0].childNodes[0].nodeValue} id: {excell.getAttribute('id')}\n")

    excel_selected = settings.getElementsByTagName('excel_selected')[0]
    excel_selected_variable[0] = excel_selected.getAttribute('id')

    # logging
    if log['load_excel_selected']:
        time = datetime.datetime.now()
        with open(logfile, 'a') as f:
            f.write(
                f"{time.hour}:{time.minute}:{time.second} load_excel_selected: id: {excel_selected.getAttribute('id')}\n")


def xml_add_person(name, tasksvariable, employees, i):
    # This function adds a new employee to the XML-file
    # It also adds the new employee to the employees array
    # and returns the current index-number.

    # Open the XML-file
    temp = []
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement

    # Set the name and default_task settings
    new = domtree.createElement('employee')
    data = domtree.createElement('name')
    data.appendChild(domtree.createTextNode(name))
    new.appendChild(data)
    data = domtree.createElement('default_task')
    data.appendChild(domtree.createTextNode('1'))
    new.appendChild(data)

    # Sets the certied setting for all available tasks
    for task in tasksvariable:
        data = domtree.createElement('task_settings')
        data.setAttribute('id', task[0])
        data2 = domtree.createElement('certified')
        data2.appendChild(domtree.createTextNode(str(task[4])))
        data.appendChild(data2)
        new.appendChild(data)
        temp.append([task[0], str(task[4])])

    # Write the XML and update the employees variable
    settings.appendChild(new)
    employees.append([name, temp, '1'])
    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")

    # Check the current index-number and returns it
    if not i:
        person_id = 0
    else:
        person_id = i + 1
    return person_id


def xml_new_task(tasksvariable, newtaskname, newtaskcolor, newtaskauto_generate, newtaskdefault_certified, newtaskschedule_length, newtaskschedule_max_times):
    # This function adds a new task to the XML-file

    # Prepare variables and opens XML
    temp = []
    task_id = str(int(tasksvariable[len(tasksvariable) - 1][0]) + 1)
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement

    # Set all the variables
    new_task = domtree.createElement('task')
    new_task.setAttribute('id', task_id)
    temp.append(task_id)
    name = domtree.createElement('name')
    name.appendChild(domtree.createTextNode(str(newtaskname.get())))
    new_task.appendChild(name)
    temp.append(str(newtaskname.get()))
    color = domtree.createElement('color')
    color.appendChild(domtree.createTextNode(newtaskcolor.get()))
    new_task.appendChild(color)
    temp.append(newtaskcolor.get())
    auto_generate = domtree.createElement('auto_generate')
    if newtaskauto_generate.get():
        auto = 'true'
    else:
        auto = 'false'
    auto_generate.appendChild(domtree.createTextNode(auto))
    new_task.appendChild(auto_generate)
    temp.append(auto)
    default_certified = domtree.createElement('default_certified')
    if newtaskdefault_certified.get():
        certified = 'true'
    else:
        certified = 'false'
    default_certified.appendChild(domtree.createTextNode(certified))
    new_task.appendChild(default_certified)
    temp.append(certified)
    schedule_length = domtree.createElement('schedule_length')
    if newtaskschedule_length.get():
        length = newtaskschedule_length.get()
    else:
        length = '60'
    schedule_length.appendChild(domtree.createTextNode(length))
    new_task.appendChild(schedule_length)
    temp.append(length)
    schedule_max_times = domtree.createElement('schedule_max_times')
    if newtaskschedule_max_times.get():
        max_times = newtaskschedule_max_times.get()
    else:
        max_times = '3'
    schedule_max_times.appendChild(domtree.createTextNode(max_times))
    new_task.appendChild(schedule_max_times)
    temp.append(max_times)
    settings.appendChild(new_task)
    domtree.appendChild(settings)
    tasksvariable.append(temp)

    # Adds the task to every employee, with default settings
    employee = settings.getElementsByTagName('employee')
    for e in employee:
        new_task2 = domtree.createElement('task_settings')
        new_task2.setAttribute('id', task_id)
        certified = domtree.createElement('certified')
        if newtaskdefault_certified.get():
            certifie = 'True'
        else:
            certifie = 'False'
        certified.appendChild(domtree.createTextNode(certifie))
        new_task2.appendChild(certified)
        e.appendChild(new_task2)

    # save XML
    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")


def xml_delete_task(tasksvariable, task, employees):
    # Deletes the task from the XML, and updates employees default task if needed

    # Load XML
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement

    # Remove the task
    tasks = settings.getElementsByTagName('task')
    for t in tasks:
        if tasksvariable[task][0] == t.getAttribute('id'):
            t.parentNode.removeChild(t)

    # Load all employees
    employee = settings.getElementsByTagName('employee')
    for e in employee:

        # Removes the task from all employees
        tasks = e.getElementsByTagName('task_settings')
        for t in tasks:
            if tasksvariable[task][0] == t.getAttribute('id'):
                t.parentNode.removeChild(t)

        # Removes the default task from all employees that has it set
        default_task = e.getElementsByTagName('default_task')[0].childNodes[0].nodeValue
        if default_task == str(tasksvariable[task][0]):
            e.getElementsByTagName('default_task')[0].childNodes[0].nodeValue = str(1)
            employees[en][2] = 1

    # Save XML
    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")


def xml_save_excel_template(ws, excell_templates, add_excel_variables):


    # Load XML
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement
    excell = domtree.createElement('excell')

    # Calculate the new excel id number, and set it in the XML
    excel_id = 0
    for e in excell_templates:
        excel_id = int(e)
    if excel_id < 10:
        excel_id = 10
    else:
        excel_id = excel_id + 1
    excell.setAttribute('id', str(excel_id))


    title = domtree.createElement('title')
    title.appendChild(domtree.createTextNode(add_excel_variables[0].get()))
    excell.appendChild(title)
    data = []
    for row in range(1, 16):
        for col in range(1, 55):
            temp = {}

            # set correct cell
            col = get_column_letter(col)
            cell = ws[f'{col}{str(row)}']
            xml_cell = domtree.createElement('cell')

            # set cell id
            xml_cell.setAttribute('id', f'{col}:{str(row)}')
            temp['id'] = f'{col}:{str(row)}'

            #set cell text if present
            text = cell.value
            if text:
                xml_cell.appendChild(domtree.createTextNode(str(text)))
            temp['text'] = str(text)

            # set font size if text and font size present
            font_size = cell.font.size
            if font_size and text:
                xml_cell.setAttribute('font_size', str(font_size))
            temp['font_size'] = str(font_size)

            # set font if font and text is present
            font = cell.font.name
            if font and text:
                xml_cell.setAttribute('font', font)
            temp['font'] = font

            # set bold if bold and text is present
            font_style_bold = cell.font.bold
            if font_style_bold and text:
                xml_cell.setAttribute('font_style_bold', str(font_style_bold))
            temp['font_style_bold'] = str(font_style_bold)

            # set italic if italic and text is present
            font_style_italic = cell.font.italic
            if font_style_italic and text:
                xml_cell.setAttribute('font_style_italic', str(font_style_italic))
            temp['font_style_italic'] = str(font_style_italic)

            # set underline type if underline and text is present
            font_style_underline = cell.font.underline
            if font_style_underline and text:
                xml_cell.setAttribute('font_style_underline', str(font_style_underline))
            temp['font_style_underline'] = str(font_style_underline)

            # set font color
            fg = cell.font.color
            if fg:
                if fg.type == 'rgb':
                    fg = fg.rgb[2:]
                    if fg:
                        xml_cell.setAttribute('fg', fg)
            temp['fg'] = fg

            border_left = cell.border.left.style
            if border_left:
                xml_cell.setAttribute('border_left', border_left)
            temp['border_left'] = border_left

            border_right = cell.border.right.style
            if border_right:
                xml_cell.setAttribute('border_right', border_right)
            temp['border_right'] = border_right

            border_top = cell.border.top.style
            if border_top:
                xml_cell.setAttribute('border_top', border_top)
            temp['border_top'] = border_top

            border_bottom = cell.border.bottom.style
            if border_bottom:
                xml_cell.setAttribute('border_bottom', border_bottom)
            temp['border_bottom'] = border_bottom

            bg = cell.fill.fgColor
            if bg:
                if bg.type == 'rgb':
                    if bg.rgb[:2] != '00':
                        xml_cell.setAttribute('bg', bg.rgb)
                        bg = True
                    else:
                        bg = False
            temp['bg'] = bg

            border_left_color = cell.border.left.color
            if border_left_color:
                if border_left_color.type == 'rgb':
                    border_left_color = border_left_color.rgb[2:]
                    if border_left_color:
                        xml_cell.setAttribute('border_left_color', border_left_color)
            temp['border_left_color'] = border_left_color

            border_right_color = cell.border.right.color
            if border_right_color:
                if border_right_color.type == 'rgb':
                    border_right_color = border_right_color.rgb[2:]
                    if border_right_color:
                        xml_cell.setAttribute('border_right_color', border_right_color)
            temp['border_right_color'] = border_right_color

            border_top_color = cell.border.top.color
            if border_top_color:
                if border_top_color.type == 'rgb':
                    border_top_color = border_top_color.rgb[2:]
                    if border_top_color:
                        xml_cell.setAttribute('border_top_color', border_top_color)
            temp['border_top_color'] = border_top_color

            border_bottom_color = cell.border.bottom.color
            if border_bottom_color:
                if border_bottom_color.type == 'rgb':
                    border_bottom_color = border_bottom_color.rgb[2:]
                    if border_bottom_color:
                        xml_cell.setAttribute('border_bottom_color', border_bottom_color)
            temp['border_bottom_color'] = border_bottom_color

            # onlo save data if one of these is set
            if text or border_left or border_right or border_top or border_bottom or bg:
                excell.appendChild(xml_cell)
                data.append(temp)

    # save xml
    settings.appendChild(excell)
    domtree.appendChild(settings)
    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")

    return data, excel_id


def xml_save_excel(excellwidgets):
    # This function saves selected excel template to xml

    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement
    excel_selected = settings.getElementsByTagName('excel_selected')[0]
    excel_selected.setAttribute('id', excellwidgets[0].get())
    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")


def delete_announcement(id):
    domtree = xml.dom.minidom.parse('settings.xml')
    settings = domtree.documentElement
    announcement = settings.getElementsByTagName('announcement')

    # Goes thru every announcement, and removes the chosen one
    for i, a in enumerate(announcement):
        if i == id:
            a.parentNode.removeChild(a)
    domtree.writexml(codecs.open('settings.xml', "w", "utf-8"), encoding="utf-8")
    